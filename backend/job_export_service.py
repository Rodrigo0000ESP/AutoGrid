import os
from datetime import datetime
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from fastapi import HTTPException

class JobExportService:
    @staticmethod
    def export_jobs(
        db: Session,
        user_id: int,
        output_dir: str = 'exports',
        filename: str = None
    ) -> str:
        """
        Export user's jobs to an Excel file.
        
        Args:
            db: Database session
            user_id: ID of the user whose jobs to export
            output_dir: Directory to save the Excel file
            filename: Custom filename (without extension)
            
        Returns:
            Path to the generated Excel file
        """
        try:
            from models import Job  # Import here to avoid circular imports
            
            # Get all jobs for the user
            jobs = db.query(Job).filter(Job.user_id == user_id).all()
            
            if not jobs:
                raise HTTPException(
                    status_code=404,
                    detail="No jobs found to export"
                )
            
            # Create output directory if it doesn't exist
            os.makedirs(output_dir, exist_ok=True)
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"jobs_export_{timestamp}.xlsx"
            
            filepath = os.path.join(output_dir, filename)
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Jobs Export"
            
            # Define columns with display names and model field mappings
            columns = [
                ('position', 'Job Title'),
                ('company', 'Company'),
                ('location', 'Location'),
                ('job_type', 'Job Type'),
                ('status', 'Status'),
                ('salary', 'Salary'),
                ('date_added', 'Date Added'),
                ('date_modified', 'Last Modified'),
                ('link', 'Job URL')
            ]
            
            # Define styles
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            header_border = Border(bottom=Side(style='medium'))
            url_font = Font(color='0000FF', underline='single')
            date_format = 'YYYY-MM-DD'
            
            # Add header row with styling
            for col_num, (_, display_name) in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num, value=display_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = header_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add data rows with proper formatting
            for row_num, job in enumerate(jobs, 2):
                for col_num, (col_name, _) in enumerate(columns, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    value = getattr(job, col_name, '')
                    
                    # Handle different data types
                    if value is None:
                        value = ''
                    elif hasattr(value, 'value'):  # Enum values
                        value = value.value
                    elif hasattr(value, 'strftime'):  # Date values
                        value = value.strftime('%Y-%m-%d')
                        cell.number_format = date_format
                    
                    # Special handling for URLs
                    if col_name == 'link' and value:
                        cell.hyperlink = value
                        cell.font = url_font
                    
                    cell.value = value
            
            # Auto-adjust column widths with some padding
            for col in ws.columns:
                max_length = 0
                column = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 4) * 1.1
                ws.column_dimensions[column].width = min(adjusted_width, 50)
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Save the workbook
            wb.save(filepath)
            
            return filepath
            
        except Exception as e:
            if isinstance(e, HTTPException):
                raise e
            raise HTTPException(
                status_code=500,
                detail=f"Error exporting jobs to Excel: {str(e)}"
            )
